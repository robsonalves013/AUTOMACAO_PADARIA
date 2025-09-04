import datetime
import pandas as pd
import json
import locale
import os
import time
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from collections import defaultdict
import smtplib
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
import uuid

# --- CÓDIGO DA API FLASK ---
from flask import Flask, jsonify, request
from flask_cors import CORS

# --- Códigos ANSI para cores ---
AMARELO = '\033[93m'
AZUL = '\033[94m'
VERDE = '\033[92m'
VERMELHO = '\033[91m'
RESET = '\033[0m'
NEGRITO = '\033[1m'

# Configura o locale para português do Brasil
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')

# Diretório para salvar os arquivos e subpastas
DIRETORIO_DADOS = "relatorios_padaria"
ASSINATURA = "Sistema desenvolvido por ROBSON ALVES"

# --- Configurações do e-mail (ATUALIZAR COM SUAS INFORMAÇÕES) ---
EMAIL_REMETENTE = 'robtechservice@outlook.com'
SENHA_APP = 'ioohmnnkugrsulss'
SENHA_MASTER = '120724'

# ----------------------------------------
# --- FUNÇÕES DE LÓGICA DO NEGÓCIO ---
# ----------------------------------------
def formatar_texto(texto, cor=RESET, estilo=RESET):
    """Função centralizada para formatar texto com cores e estilos."""
    return f"{estilo}{cor}{texto}{RESET}"

def verificar_diretorio():
    """Cria o diretório de dados se ele não existir."""
    if not os.path.exists(DIRETORIO_DADOS):
        os.makedirs(DIRETORIO_DADOS)

def carregar_dados():
    """Carrega os dados de um arquivo JSON. Se não existir, cria um com dados de exemplo."""
    verificar_diretorio()
    caminho_arquivo = os.path.join(DIRETORIO_DADOS, 'dados_padaria.json')
    try:
        with open(caminho_arquivo, 'r') as f:
            dados = json.load(f)
            # Certifica-se de que o histórico de estoque existe
            if 'historico_estoque' not in dados:
                dados['historico_estoque'] = []
            return dados['receitas'], dados['despesas'], dados['estoque'], dados['historico_estoque']
    except FileNotFoundError:
        print(formatar_texto("Arquivo de dados não encontrado. Criando um novo com dados de exemplo...", cor=AMARELO))
        receitas_exemplo = []
        despesas_exemplo = []
        estoque_exemplo = {
            '123456789': {'descricao': 'pão francês', 'quantidade': 100, 'valor_unitario': 0.50, 'categoria': 'Pães'},
            '987654321': {'descricao': 'pão de queijo', 'quantidade': 50, 'valor_unitario': 3.00, 'categoria': 'Pães'},
            '456789123': {'descricao': 'bolo de chocolate', 'quantidade': 10, 'valor_unitario': 25.00, 'categoria': 'Doces'},
            '789123456': {'descricao': 'torta de limão', 'quantidade': 8, 'valor_unitario': 30.00, 'categoria': 'Doces'},
            '111222333': {'descricao': 'café expresso', 'quantidade': 100, 'valor_unitario': 4.50, 'categoria': 'Bebidas'},
            '444555666': {'descricao': 'suco de laranja', 'quantidade': 20, 'valor_unitario': 6.00, 'categoria': 'Bebidas'}
        }
        historico_estoque_exemplo = []
        salvar_dados(receitas_exemplo, despesas_exemplo, estoque_exemplo, historico_estoque_exemplo)
        return receitas_exemplo, despesas_exemplo, estoque_exemplo, historico_estoque_exemplo

def salvar_dados(receitas, despesas, estoque, historico_estoque):
    """Salva os dados em um arquivo JSON e exibe uma mensagem de sucesso."""
    verificar_diretorio()
    caminho_arquivo = os.path.join(DIRETORIO_DADOS, 'dados_padaria.json')
    dados = {
        'receitas': receitas,
        'despesas': despesas,
        'estoque': estoque,
        'historico_estoque': historico_estoque
    }
    try:
        with open(caminho_arquivo, 'w') as f:
            json.dump(dados, f, indent=4)
        print(formatar_texto("Dados salvos com sucesso!", cor=VERDE))
    except Exception as e:
        print(formatar_texto(f"Erro ao salvar os dados: {e}", cor=VERMELHO))

def adicionar_produto_estoque(estoque):
    """Permite adicionar ou atualizar um produto no estoque via terminal."""
    codigo = input("Digite o código do produto (código de barras): ")
    descricao = input("Digite a descrição do produto: ")
    try:
        quantidade = int(input("Digite a quantidade a ser adicionada: "))
        valor_unitario = float(input("Digite o valor unitário: "))
    except ValueError:
        print(formatar_texto("Entrada inválida para quantidade ou valor. Tente novamente.", cor=VERMELHO))
        return
    categoria = input("Digite a categoria do produto: ")

    if codigo in estoque:
        estoque[codigo]['quantidade'] += quantidade
        print(formatar_texto("Produto atualizado com sucesso!", cor=VERDE))
    else:
        estoque[codigo] = {
            'descricao': descricao.lower(),
            'quantidade': quantidade,
            'valor_unitario': valor_unitario,
            'categoria': categoria
        }
        print(formatar_texto("Produto adicionado com sucesso!", cor=VERDE))

def registrar_venda_direta(receitas, estoque):
    """Registra uma venda direta a partir do terminal."""
    codigo = input("Digite o código do produto vendido: ")
    try:
        quantidade = int(input("Digite a quantidade vendida: "))
    except ValueError:
        print(formatar_texto("Quantidade inválida. Tente novamente.", cor=VERMELHO))
        return

    if codigo not in estoque or estoque[codigo]['quantidade'] < quantidade:
        print(formatar_texto("Estoque insuficiente ou produto não encontrado.", cor=VERMELHO))
        return

    forma_pagamento = input("Digite a forma de pagamento (ex: 'Cartão', 'Dinheiro', 'PIX'): ")
    
    valor_unitario = estoque[codigo]['valor_unitario']
    valor_total = quantidade * valor_unitario
    
    estoque[codigo]['quantidade'] -= quantidade
    
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'id_transacao': uuid.uuid4().hex,
        'itens': [{
            'codigo_produto': codigo,
            'quantidade': quantidade,
            'valor_unitario': valor_unitario,
            'descricao': estoque[codigo]['descricao']
        }],
        'descricao': f"Venda de {quantidade} und. de {estoque[codigo]['descricao']}",
        'valor': valor_total,
        'data': agora,
        'tipo': 'receita',
        'forma_pagamento': forma_pagamento
    })
    
    print(formatar_texto(f"Venda de {valor_total:,.2f} registrada com sucesso!", cor=VERDE))

def registrar_venda_delivery(receitas, estoque):
    """Registra uma venda por delivery a partir do terminal."""
    codigo = input("Digite o código do produto vendido: ")
    try:
        quantidade = int(input("Digite a quantidade vendida: "))
    except ValueError:
        print(formatar_texto("Quantidade inválida. Tente novamente.", cor=VERMELHO))
        return

    if codigo not in estoque or estoque[codigo]['quantidade'] < quantidade:
        print(formatar_texto("Estoque insuficiente ou produto não encontrado.", cor=VERMELHO))
        return
        
    plataforma = input("Digite a plataforma de delivery (ex: 'iFood', 'Rappi'): ")
    
    valor_total = 0 # Valor para delivery é zero
    
    estoque[codigo]['quantidade'] -= quantidade
    
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'id_transacao': uuid.uuid4().hex,
        'itens': [{
            'codigo_produto': codigo,
            'quantidade': quantidade,
            'valor_unitario': estoque[codigo]['valor_unitario'],
            'descricao': estoque[codigo]['descricao']
        }],
        'descricao': f"Venda de {quantidade} und. de {estoque[codigo]['descricao']} (Delivery) - SEM VALOR",
        'valor': valor_total,
        'data': agora,
        'tipo': 'receita',
        'plataforma_delivery': plataforma
    })
    
    print(formatar_texto(f"Venda por delivery de {valor_total:,.2f} registrada com sucesso!", cor=VERDE))

def adicionar_receita(receitas):
    """Adiciona uma receita manual ao fluxo de caixa."""
    descricao = input("Descrição da receita: ")
    try:
        valor = float(input("Valor da receita: "))
    except ValueError:
        print(formatar_texto("Valor inválido. Tente novamente.", cor=VERMELHO))
        return

    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'descricao': descricao,
        'valor': valor,
        'data': agora,
        'tipo': 'receita',
        'forma_pagamento': 'Manual'
    })
    print(formatar_texto("Receita adicionada com sucesso!", cor=VERDE))

def adicionar_despesa(despesas):
    """Adiciona uma despesa manual ao fluxo de caixa."""
    descricao = input("Descrição da despesa: ")
    try:
        valor = float(input("Valor da despesa: "))
    except ValueError:
        print(formatar_texto("Valor inválido. Tente novamente.", cor=VERMELHO))
        return

    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    despesas.append({
        'descricao': descricao,
        'valor': valor,
        'data': agora,
        'tipo': 'despesa'
    })
    print(formatar_texto("Despesa adicionada com sucesso!", cor=VERDE))

# ----------------------------------------
# --- FUNÇÕES DE API (FLASK) ---
# ----------------------------------------
app = Flask(__name__)
CORS(app)

@app.route('/estoque', methods=['GET'])
def get_estoque():
    """Endpoint para a API web buscar o estoque atual."""
    receitas, despesas, estoque, historico = carregar_dados()
    return jsonify(estoque)

@app.route('/estoque/add', methods=['POST'])
def add_produto():
    """Endpoint para adicionar/atualizar um produto via API."""
    data = request.json
    codigo = data.get('codigo')
    descricao = data.get('descricao')
    quantidade = data.get('quantidade')
    valor_unitario = data.get('valor_unitario')
    categoria = data.get('categoria')

    if not all([codigo, descricao, quantidade, valor_unitario]):
        return jsonify({"error": "Dados incompletos"}), 400

    receitas, despesas, estoque, historico = carregar_dados()
    if codigo in estoque:
        estoque[codigo]['quantidade'] += quantidade
        estoque[codigo]['valor_unitario'] = valor_unitario
        estoque[codigo]['categoria'] = categoria
        message = "Produto atualizado com sucesso!"
    else:
        estoque[codigo] = {
            'descricao': descricao.lower(),
            'quantidade': quantidade,
            'valor_unitario': valor_unitario,
            'categoria': categoria
        }
        message = "Produto adicionado com sucesso!"

    salvar_dados(receitas, despesas, estoque, historico)
    return jsonify({"message": message})
    
@app.route('/estoque/editar', methods=['POST'])
def editar_estoque_api():
    """Endpoint para editar a quantidade de um produto com senha master."""
    data = request.json
    codigo = data.get('codigo')
    nova_quantidade = data.get('nova_quantidade')
    senha = data.get('senha_master')
    
    if not all([codigo, nova_quantidade, senha]):
        return jsonify({"error": "Dados incompletos"}), 400
    
    receitas, despesas, estoque, historico = carregar_dados()
    
    if senha != SENHA_MASTER:
        return jsonify({"error": "Senha master incorreta. Acesso negado."}), 401
    
    if codigo not in estoque:
        return jsonify({"error": "Produto não encontrado."}), 404
        
    quantidade_anterior = estoque[codigo]['quantidade']
    estoque[codigo]['quantidade'] = nova_quantidade
    
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    historico.append({
        'data_hora': agora,
        'codigo': codigo,
        'descricao': estoque[codigo]['descricao'],
        'quantidade_anterior': quantidade_anterior,
        'nova_quantidade': nova_quantidade
    })
    
    salvar_dados(receitas, despesas, estoque, historico)
    return jsonify({"message": "Estoque atualizado com sucesso!", "historico": historico})


@app.route('/vendas/direta', methods=['POST'])
def post_venda_direta():
    """
    Endpoint para registrar uma venda direta com múltiplos itens via API.
    O payload deve conter a lista de itens do carrinho.
    """
    data = request.json
    carrinho = data.get('carrinho')
    forma_pagamento = data.get('forma_pagamento')

    if not carrinho or not forma_pagamento:
        return jsonify({"error": "Dados de venda incompletos"}), 400

    receitas, despesas, estoque, historico = carregar_dados()
    valor_total = 0
    itens_vendidos = []

    # 1. Verifica o estoque de todos os itens antes de processar a venda
    for item in carrinho:
        codigo = item.get('codigo')
        quantidade = item.get('quantidade')
        
        if codigo not in estoque or estoque[codigo]['quantidade'] < quantidade:
            return jsonify({"error": f"Estoque insuficiente para o item {codigo} ou produto não encontrado."}), 400
        
        valor_unitario = estoque[codigo]['valor_unitario']
        valor_total += quantidade * valor_unitario
        itens_vendidos.append({
            'codigo_produto': codigo,
            'quantidade': quantidade,
            'valor_unitario': valor_unitario,
            'descricao': estoque[codigo]['descricao']
        })

    # 2. Processa a venda e atualiza o estoque e as receitas
    for item in carrinho:
        codigo = item.get('codigo')
        quantidade = item.get('quantidade')
        estoque[codigo]['quantidade'] -= quantidade

    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'id_transacao': uuid.uuid4().hex,
        'itens': itens_vendidos,
        'descricao': f"Venda de múltiplos itens",
        'valor': valor_total,
        'data': agora,
        'tipo': 'receita',
        'forma_pagamento': forma_pagamento
    })

    salvar_dados(receitas, despesas, estoque, historico)
    return jsonify({"message": "Venda direta registrada com sucesso!", "valor_total": valor_total})

@app.route('/vendas/delivery', methods=['POST'])
def post_venda_delivery():
    """Endpoint para registrar uma venda delivery via API."""
    data = request.json
    codigo = data.get('codigo')
    quantidade = data.get('quantidade')
    plataforma = data.get('plataforma')

    if not all([codigo, quantidade, plataforma]):
        return jsonify({"error": "Dados de venda incompletos"}), 400

    receitas, despesas, estoque, historico = carregar_dados()
    if codigo not in estoque or estoque[codigo]['quantidade'] < quantidade:
        return jsonify({"error": "Estoque insuficiente ou produto não encontrado"}), 400

    # Valor de vendas delivery sempre é 0, pois a cobrança é feita pela plataforma
    valor_total = 0
    estoque[codigo]['quantidade'] -= quantidade

    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'id_transacao': uuid.uuid4().hex,
        'itens': [{
            'codigo_produto': codigo,
            'quantidade': quantidade,
            'valor_unitario': estoque[codigo]['valor_unitario'],
            'descricao': estoque[codigo]['descricao']
        }],
        'descricao': f"Venda de {quantidade} und. de {estoque[codigo]['descricao']} (Delivery)",
        'valor': valor_total,
        'data': agora,
        'tipo': 'receita',
        'plataforma_delivery': plataforma
    })

    salvar_dados(receitas, despesas, estoque, historico)
    return jsonify({"message": "Venda delivery registrada com sucesso!"})

@app.route('/vendas/cancelar', methods=['POST'])
def post_venda_cancelar():
    """
    Endpoint para cancelar uma venda feita pela API, usando o ID de transação e uma senha master. Agora lida com vendas de múltiplos itens.
    """
    data = request.json
    id_transacao = data.get('id_transacao')
    senha = data.get('senha_master')
    if not all([id_transacao, senha]):
        return jsonify({"error": "ID de transação ou senha não fornecidos"}), 400
    
    if senha != SENHA_MASTER:
        return jsonify({"error": "Senha incorreta. Acesso negado."}), 401
        
    receitas, despesas, estoque, historico = carregar_dados()
    venda_encontrada = None
    for venda in receitas:
        if venda.get('id_transacao') == id_transacao:
            venda_encontrada = venda
            break
            
    if not venda_encontrada:
        return jsonify({"error": "Venda não encontrada"}), 404
        
    if 'CANCELADA' in venda_encontrada.get('descricao', ''):
        return jsonify({"error": "Esta venda já foi cancelada."}), 400

    try:
        itens_vendidos = venda_encontrada.get('itens', [])
        if not itens_vendidos:
            return jsonify({"error": "Dados de itens da venda incompletos para cancelamento"}), 500
        
        # Restaura o estoque para todos os itens na venda
        for item in itens_vendidos:
            codigo_produto = item.get('codigo_produto')
            quantidade = item.get('quantidade')
            if codigo_produto in estoque and quantidade is not None:
                estoque[codigo_produto]['quantidade'] += quantidade
            else:
                return jsonify({"error": f"Erro: Item {codigo_produto} não encontrado ou dados incompletos para restauração de estoque."}), 500
                
        # Altera a descrição e valor da venda para indicar que foi cancelada
        venda_encontrada['descricao'] = f"{venda_encontrada.get('descricao')} - CANCELADA"
        venda_encontrada['valor'] = 0
        
        salvar_dados(receitas, despesas, estoque, historico)
        return jsonify({"message": f"Venda com ID {id_transacao} cancelada com sucesso. Estoque e dados atualizados."})
    except Exception as e:
        return jsonify({"error": f"Erro ao cancelar a venda: {e}"}), 500

@app.route('/fluxo_caixa/receita', methods=['POST'])
def add_receita_api():
    """Endpoint para adicionar uma receita manual via API."""
    data = request.json
    descricao = data.get('descricao')
    valor = data.get('valor')
    if not all([descricao, valor]):
        return jsonify({"error": "Dados de receita incompletos"}), 400

    receitas, despesas, estoque, historico = carregar_dados()
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'descricao': descricao,
        'valor': valor,
        'data': agora,
        'tipo': 'receita',
        'forma_pagamento': 'Manual'
    })
    salvar_dados(receitas, despesas, estoque, historico)
    return jsonify({"message": "Receita adicionada com sucesso!"})

@app.route('/fluxo_caixa/despesa', methods=['POST'])
def add_despesa_api():
    """Endpoint para adicionar uma despesa via API."""
    data = request.json
    descricao = data.get('descricao')
    valor = data.get('valor')
    if not all([descricao, valor]):
        return jsonify({"error": "Dados de despesa incompletos"}), 400

    receitas, despesas, estoque, historico = carregar_dados()
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    despesas.append({
        'descricao': descricao,
        'valor': valor,
        'data': agora,
        'tipo': 'despesa'
    })
    salvar_dados(receitas, despesas, estoque, historico)
    return jsonify({"message": "Despesa adicionada com sucesso!"})

@app.route('/vendas/diarias', methods=['GET'])
def get_vendas_diarias():
    """Endpoint para a API web buscar as vendas do dia."""
    receitas, _, _, _ = carregar_dados()
    hoje = datetime.date.today()
    vendas_hoje = [
        item for item in receitas
        if item.get('tipo') == 'receita' and datetime.datetime.strptime(item['data'], "%d/%m/%Y %H:%M:%S").date() == hoje
    ]
    return jsonify(vendas_hoje)

@app.route('/historico_estoque', methods=['GET'])
def get_historico_estoque():
    """Endpoint para a API web buscar o histórico de alterações no estoque."""
    _, _, _, historico = carregar_dados()
    return jsonify(historico)


# ----------------------------------------
# --- FUNÇÕES DE RELATÓRIOS E TERMINAL ---
# ----------------------------------------
def limpar_tela():
    """ Limpa o console de uma forma mais segura usando códigos ANSI. Substitui a versão anterior baseada em os.system() para evitar travamentos. """
    sys.stdout.write('\033[H\033[J')

def mostrar_logo_inicial():
    """Exibe a mensagem de abertura do script."""
    print("...")

def menu_principal():
    """Exibe o menu principal e processa a escolha do usuário."""
    receitas, despesas, estoque, historico = carregar_dados()

    while True:
        limpar_tela()
        mostrar_logo_inicial()
        print(formatar_texto("--- MENU PRINCIPAL ---", cor=VERDE, estilo=NEGRITO))
        print("1. Gerar Relatórios")
        print("2. Adicionar Receita")
        print("3. Adicionar Despesa")
        print("4. Adicionar Produto ao Estoque")
        print("5. Registrar Venda Direta")
        print("6. Registrar Venda Delivery")
        print("7. Verificar Estoque e Enviar Alerta")
        print(formatar_texto("8. Sair", cor=VERMELHO, estilo=NEGRITO))
        
        escolha = input("Escolha uma opção: ")

        if escolha == '1':
            menu_gerar_relatorio(receitas, despesas, estoque)
        elif escolha == '2':
            adicionar_receita(receitas)
        elif escolha == '3':
            adicionar_despesa(despesas)
        elif escolha == '4':
            adicionar_produto_estoque(estoque)
        elif escolha == '5':
            registrar_venda_direta(receitas, estoque)
        elif escolha == '6':
            registrar_venda_delivery(receitas, estoque)
        elif escolha == '7':
            verificar_e_enviar_alerta_estoque(estoque)
        elif escolha == '8':
            print("Salvando dados...")
            salvar_dados(receitas, despesas, estoque, historico)
            print("Encerrando o programa.")
            break
        else:
            print(formatar_texto("Opção inválida. Tente novamente.", cor=VERMELHO))
        
        input(formatar_texto("\nPressione Enter para continuar...", cor=AZUL, estilo=NEGRITO))
        limpar_tela()

if __name__ == '__main__':
    # Cria o diretório de dados se ele não existir
    verificar_diretorio()
    # Inicia o servidor Flask em uma thread separada para não bloquear o terminal
    import threading
    threading.Thread(target=lambda: app.run(debug=False, use_reloader=False)).start()
    
    print("Servidor web iniciado em http://127.0.0.1:5000")
    menu_principal()