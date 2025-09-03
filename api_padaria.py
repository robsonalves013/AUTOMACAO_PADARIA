import datetime
import pandas as pd
import json
import locale
import os
from flask import Flask, request, jsonify, send_file
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import smtplib
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders

# --- Configurações do Flask e dados ---
app = Flask(__name__)
receitas, despesas, estoque = {}, {}, {}  # Variáveis globais para armazenar os dados

# Diretório para salvar os arquivos
DIRETORIO_DADOS = "relatorios_padaria"
ASSINATURA = "Sistema desenvolvido por ROBSON ALVES"

# --- Configurações do e-mail (ATUALIZAR COM SUAS INFORMAÇÕES) ---
EMAIL_REMETENTE = 'robtechservice@outlook.com'
SENHA_APP = 'ioohmnnkugrsulss'

# Configura o locale para português do Brasil
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')

def verificar_diretorio():
    """Cria o diretório de dados se ele não existir."""
    if not os.path.exists(DIRETORIO_DADOS):
        os.makedirs(DIRETORIO_DADOS)

def carregar_dados():
    """Carrega os dados de um arquivo JSON. Se não existir, cria um com dados de exemplo."""
    verificar_diretorio()
    caminho_arquivo = os.path.join(DIRETORIO_DADOS, 'dados_padaria.json')
    try:
        with open(caminho_arquivo, 'r') as f:
            dados = json.load(f)
            # Garante que as chaves de estoque tenham 'descricao'
            for key in dados.get('estoque', {}).keys():
                if 'descricao' not in dados['estoque'][key]:
                    dados['estoque'][key]['descricao'] = key
            return dados.get('receitas', []), dados.get('despesas', []), dados.get('estoque', {})
    except FileNotFoundError:
        receitas_exemplo = []
        despesas_exemplo = []
        estoque_exemplo = {
            '123456789': {'descricao': 'pão francês', 'quantidade': 100, 'valor_unitario': 0.50, 'categoria': 'Pães'},
            '987654321': {'descricao': 'pão de queijo', 'quantidade': 50, 'valor_unitario': 3.00, 'categoria': 'Pães'},
            '456789123': {'descricao': 'bolo de chocolate', 'quantidade': 10, 'valor_unitario': 25.00, 'categoria': 'Doces'},
        }
        salvar_dados(receitas_exemplo, despesas_exemplo, estoque_exemplo)
        return receitas_exemplo, despesas_exemplo, estoque_exemplo

def salvar_dados(receitas, despesas, estoque):
    """Salva os dados em um arquivo JSON."""
    verificar_diretorio()
    caminho_arquivo = os.path.join(DIRETORIO_DADOS, 'dados_padaria.json')
    dados = {
        'receitas': receitas,
        'despesas': despesas,
        'estoque': estoque
    }
    with open(caminho_arquivo, 'w') as f:
        json.dump(dados, f, indent=4)
    print("Dados salvos com sucesso.")

def criar_subdiretorio(subpasta):
    """Cria um subdiretório dentro do diretório de dados se ele não existir."""
    caminho_completo = os.path.join(DIRETORIO_DADOS, subpasta)
    if not os.path.exists(caminho_completo):
        os.makedirs(caminho_completo)
    return caminho_completo

def formatar_planilha_excel(caminho_arquivo):
    """Formata as planilhas do Excel para melhor visualização."""
    try:
        workbook = load_workbook(caminho_arquivo)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            header = [cell.value for cell in worksheet[1]]
            if 'valor' in header or 'Valor' in header:
                try:
                    col_index = header.index('valor') + 1
                except ValueError:
                    col_index = header.index('Valor') + 1
                for row in worksheet.iter_rows(min_row=2, max_col=col_index, max_row=worksheet.max_row):
                    cell = row[col_index - 1]
                    cell.number_format = 'R$ #,##0.00'
        workbook.save(caminho_arquivo)
    except Exception as e:
        print(f"Erro ao formatar o arquivo Excel: {e}")

def adicionar_assinatura_excel(caminho_arquivo):
    """Adiciona a assinatura de desenvolvimento no rodapé de cada planilha."""
    try:
        workbook = load_workbook(caminho_arquivo)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            worksheet.oddFooter.right.text = ASSINATURA
            worksheet.oddFooter.center.text = "Página &[Page] de &N"
        workbook.save(caminho_arquivo)
    except Exception as e:
        print(f"Aviso: Não foi possível adicionar a assinatura no relatório. Erro: {e}")

def enviar_email_com_anexo(assunto, corpo, destinatario, caminho_anexo=None):
    """Envia um e-mail com anexo."""
    remetente_email = EMAIL_REMETENTE
    remetente_senha = SENHA_APP
    msg = MIMEMultipart()
    msg['From'] = remetente_email
    msg['To'] = destinatario
    msg['Subject'] = assunto
    msg.attach(MIMEText(corpo, 'plain'))
    if caminho_anexo and os.path.exists(caminho_anexo):
        try:
            with open(caminho_anexo, 'rb') as anexo:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(anexo.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(caminho_anexo)}"')
            msg.attach(part)
        except Exception as e:
            print(f"Aviso: Erro ao anexar o arquivo: {e}")
    try:
        server = smtplib.SMTP_SSL('smtp.outlook.com', 465)
        server.login(remetente_email, remetente_senha)
        server.sendmail(remetente_email, destinatario, msg.as_string())
        server.quit()
        print(f"E-mail com o relatório enviado para {destinatario} com sucesso.")
        return True
    except Exception as e:
        print(f"Erro ao enviar o e-mail: {e}")
        return False

def gerar_relatorios_estoque_excel():
    """Gera o relatório de estoque em formato de planilha e retorna o caminho."""
    subpasta = "Estoque"
    caminho_subpasta = criar_subdiretorio(subpasta)
    dados_estoque = [{'Código de Barras': codigo, 'Produto': dados.get('descricao', codigo).capitalize(), 'Quantidade': dados['quantidade'], 'Valor Unitário': dados['valor_unitario'], 'Categoria': dados.get('categoria', 'N/A')} for codigo, dados in estoque.items()]
    df_estoque = pd.DataFrame(dados_estoque)
    caminho_arquivo = os.path.join(caminho_subpasta, 'Relatorio_de_Estoque.xlsx')
    if not df_estoque.empty:
        df_estoque.to_excel(caminho_arquivo, index=False)
        formatar_planilha_excel(caminho_arquivo)
        adicionar_assinatura_excel(caminho_arquivo)
        return caminho_arquivo
    return None

def gerar_relatorio_fluxo_caixa_excel(mes=None, ano=None):
    """Gera o relatório de fluxo de caixa em formato de planilha e retorna o caminho."""
    subpasta = "Fluxo_de_Caixa"
    caminho_subpasta = criar_subdiretorio(subpasta)
    if mes and ano:
        receitas_filtradas = [d for d in receitas if datetime.datetime.strptime(d['data'], "%d/%m/%Y %H:%M:%S").month == mes and datetime.datetime.strptime(d['data'], "%d/%m/%Y %H:%M:%S").year == ano]
        despesas_filtradas = [d for d in despesas if datetime.datetime.strptime(d['data'], "%d/%m/%Y %H:%M:%S").month == mes and datetime.datetime.strptime(d['data'], "%d/%m/%Y %H:%M:%S").year == ano]
        titulo_relatorio = f"Relatorio_de_Fluxo_de_Caixa_{mes:02d}-{ano}.xlsx"
    else:
        receitas_filtradas = receitas
        despesas_filtradas = despesas
        titulo_relatorio = "Relatorio_de_Fluxo_de_Caixa_Geral.xlsx"
    df_receitas = pd.DataFrame(receitas_filtradas)
    df_despesas = pd.DataFrame(despesas_filtradas)
    total_receitas = df_receitas['valor'].sum() if not df_receitas.empty else 0
    total_despesas = df_despesas['valor'].sum() if not df_despesas.empty else 0
    saldo = total_receitas - total_despesas
    resumo_dados = {'Métrica': ['Total de Receitas', 'Total de Despesas', 'Saldo em Caixa'], 'Valor': [total_receitas, total_despesas, saldo]}
    df_resumo = pd.DataFrame(resumo_dados)
    caminho_arquivo = os.path.join(caminho_subpasta, titulo_relatorio)
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
        df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
        if not df_receitas.empty:
            df_receitas.to_excel(writer, sheet_name='Receitas', index=False)
        if not df_despesas.empty:
            df_despesas.to_excel(writer, sheet_name='Despesas', index=False)
    formatar_planilha_excel(caminho_arquivo)
    adicionar_assinatura_excel(caminho_arquivo)
    return caminho_arquivo

# Carrega os dados na inicialização da aplicação
@app.before_first_request
def setup():
    global receitas, despesas, estoque
    receitas, despesas, estoque = carregar_dados()

# --- Rotas da API ---

@app.route('/')
def home():
    """Rota de teste para verificar se a API está online."""
    return jsonify({"message": "API da Padaria Online!"})

# --- Rotas de Estoque ---
@app.route('/estoque', methods=['GET'])
def listar_estoque():
    """Retorna a lista completa do estoque."""
    return jsonify(estoque)

@app.route('/estoque/add_unico', methods=['POST'])
def adicionar_produto_unico():
    """Adiciona ou atualiza um único produto no estoque."""
    data = request.json
    if not data or 'codigo' not in data or 'descricao' not in data or 'quantidade' not in data or 'valor_unitario' not in data:
        return jsonify({"error": "Dados incompletos"}), 400
    codigo = data['codigo']
    quantidade = data['quantidade']
    descricao = data['descricao']
    valor_unitario = data['valor_unitario']
    categoria = data.get('categoria', 'N/A')
    if codigo in estoque:
        estoque[codigo]['quantidade'] += quantidade
    else:
        estoque[codigo] = {'descricao': descricao.lower(), 'quantidade': quantidade, 'valor_unitario': valor_unitario, 'categoria': categoria}
    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": f"Produto '{descricao}' adicionado/atualizado com sucesso!"}), 201

@app.route('/estoque/add_multiplos', methods=['POST'])
def adicionar_multiplos_produtos_api():
    """Adiciona múltiplos produtos ao estoque de uma só vez."""
    data = request.json
    if not data or 'produtos' not in data:
        return jsonify({"error": "Dados incompletos. 'produtos' é obrigatório."}), 400
    
    for prod in data['produtos']:
        if not all(k in prod for k in ('codigo', 'descricao', 'quantidade', 'valor_unitario')):
            return jsonify({"error": "Dados de produto incompletos"}), 400
        codigo = prod['codigo']
        quantidade = prod['quantidade']
        descricao = prod['descricao']
        valor_unitario = prod['valor_unitario']
        categoria = prod.get('categoria', 'N/A')
        
        if codigo in estoque:
            estoque[codigo]['quantidade'] += quantidade
            estoque[codigo]['valor_unitario'] = valor_unitario
            estoque[codigo]['categoria'] = categoria
            estoque[codigo]['descricao'] = descricao.lower()
        else:
            estoque[codigo] = {'descricao': descricao.lower(), 'quantidade': quantidade, 'valor_unitario': valor_unitario, 'categoria': categoria}
    
    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": "Produtos adicionados/atualizados com sucesso!"}), 201

@app.route('/estoque/mudar_categoria', methods=['POST'])
def mudar_categoria_produto():
    """Muda a categoria de um produto no estoque."""
    data = request.json
    if not data or 'codigo' not in data or 'nova_categoria' not in data:
        return jsonify({"error": "Dados incompletos"}), 400
    codigo = data['codigo']
    nova_categoria = data['nova_categoria']
    if codigo not in estoque:
        return jsonify({"error": "Produto não encontrado"}), 404
    estoque[codigo]['categoria'] = nova_categoria
    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": f"Categoria de '{estoque[codigo]['descricao']}' alterada para '{nova_categoria}' com sucesso."}), 200

# --- Rotas de Fluxo de Caixa ---
@app.route('/fluxo_caixa/resumo', methods=['GET'])
def ver_resumo_fluxo_caixa():
    """Retorna o resumo atual do fluxo de caixa."""
    total_receitas = sum(item['valor'] for item in receitas)
    total_despesas = sum(item['valor'] for item in despesas)
    saldo = total_receitas - total_despesas
    return jsonify({
        "total_receitas": total_receitas,
        "total_despesas": total_despesas,
        "saldo": saldo,
        "receitas_detalhes": receitas,
        "despesas_detalhes": despesas
    })

@app.route('/fluxo_caixa/receita', methods=['POST'])
def adicionar_receita():
    """Adiciona uma receita manual ao fluxo de caixa."""
    data = request.json
    if not data or 'valor' not in data or 'descricao' not in data:
        return jsonify({"error": "Dados incompletos"}), 400
    valor = data['valor']
    descricao = data['descricao']
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({'descricao': descricao, 'valor': valor, 'data': agora, 'tipo': 'receita', 'forma_pagamento': 'Manual'})
    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": f"Receita de {valor} adicionada com sucesso."}), 201

@app.route('/fluxo_caixa/despesa', methods=['POST'])
def adicionar_despesa():
    """Adiciona uma despesa ao fluxo de caixa."""
    data = request.json
    if not data or 'valor' not in data or 'descricao' not in data:
        return jsonify({"error": "Dados incompletos"}), 400
    valor = data['valor']
    descricao = data['descricao']
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    despesas.append({'descricao': descricao, 'valor': valor, 'data': agora, 'tipo': 'despesa'})
    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": f"Despesa de {valor} adicionada com sucesso."}), 201

# --- Rotas de Vendas ---
@app.route('/vendas/direta', methods=['POST'])
def registrar_venda_direta():
    """Registra uma venda direta (com registro de receita) no sistema."""
    data = request.json
    if not data or 'codigo' not in data or 'quantidade' not in data or 'forma_pagamento' not in data:
        return jsonify({"error": "Dados de venda incompletos"}), 400
    codigo = data['codigo']
    quantidade = data['quantidade']
    forma_pagamento = data['forma_pagamento']
    if codigo not in estoque:
        return jsonify({"error": "Produto não encontrado"}), 404
    if estoque[codigo]['quantidade'] < quantidade:
        return jsonify({"error": "Estoque insuficiente"}), 400
    estoque[codigo]['quantidade'] -= quantidade
    valor_venda = quantidade * estoque[codigo]['valor_unitario']
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'descricao': f"Venda de {quantidade} und. de {estoque[codigo]['descricao']}",
        'valor': valor_venda, 'data': agora, 'tipo': 'receita', 'forma_pagamento': forma_pagamento
    })
    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": "Venda registrada com sucesso!", "total": valor_venda}), 200

@app.route('/vendas/delivery', methods=['POST'])
def registrar_venda_delivery():
    """Registra uma venda de delivery (sem receita) no sistema."""
    data = request.json
    if not data or 'codigo' not in data or 'quantidade' not in data or 'plataforma' not in data:
        return jsonify({"error": "Dados de venda incompletos"}), 400
    codigo = data['codigo']
    quantidade = data['quantidade']
    plataforma = data['plataforma']
    if codigo not in estoque:
        return jsonify({"error": "Produto não encontrado"}), 404
    if estoque[codigo]['quantidade'] < quantidade:
        return jsonify({"error": "Estoque insuficiente"}), 400
    estoque[codigo]['quantidade'] -= quantidade
    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": f"Venda de {quantidade} und. de {estoque[codigo]['descricao']} (Delivery) registrada no estoque."}), 200

# --- Rotas de Relatórios ---
@app.route('/relatorios/estoque', methods=['GET'])
def get_relatorio_estoque():
    """Gera e retorna o relatório de estoque em formato de arquivo Excel."""
    caminho_arquivo = gerar_relatorios_estoque_excel()
    if not caminho_arquivo:
        return jsonify({"error": "Estoque vazio. Não foi possível gerar o relatório."}), 404
    return send_file(caminho_arquivo, as_attachment=True, download_name=os.path.basename(caminho_arquivo))

@app.route('/relatorios/fluxo_caixa', methods=['GET'])
def get_relatorio_fluxo_caixa():
    """Gera e retorna o relatório de fluxo de caixa em formato de arquivo Excel.
    Pode ser filtrado por mês e ano. Ex: /relatorios/fluxo_caixa?mes=9&ano=2024
    """
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)
    caminho_arquivo = gerar_relatorio_fluxo_caixa_excel(mes, ano)
    return send_file(caminho_arquivo, as_attachment=True, download_name=os.path.basename(caminho_arquivo))

@app.route('/relatorios/vendas_diarias', methods=['GET'])
def get_relatorio_vendas_diarias():
    """Gera e retorna o relatório de vendas diárias em formato de arquivo Excel.
    Ex: /relatorios/vendas_diarias?data=01-09-2024
    """
    data_relatorio_str = request.args.get('data')
    if not data_relatorio_str:
        return jsonify({"error": "Parâmetro 'data' (DD-MM-AAAA) é obrigatório."}), 400
    
    try:
        data_relatorio_obj = datetime.datetime.strptime(data_relatorio_str, "%d-%m-%Y").date()
    except ValueError:
        return jsonify({"error": "Formato de data inválido. Use DD-MM-AAAA."}), 400
    
    vendas_dia = [item for item in receitas if item.get('tipo') == 'receita' and datetime.datetime.strptime(item['data'], "%d/%m/%Y %H:%M:%S").date() == data_relatorio_obj]
    if not vendas_dia:
        return jsonify({"message": f"Não há vendas registradas para a data {data_relatorio_str}."}), 404

    df_vendas_dia = pd.DataFrame(vendas_dia)
    relatorio_agregado = df_vendas_dia.groupby('forma_pagamento')['valor'].sum().reset_index()
    total_vendas = relatorio_agregado['valor'].sum()
    relatorio_agregado.loc[len(relatorio_agregado)] = ['Total Geral', total_vendas]
    
    subpasta = "Vendas_Diarias"
    caminho_subpasta = criar_subdiretorio(subpasta)
    caminho_arquivo = os.path.join(caminho_subpasta, f"Relatorio_de_Vendas_Diarias_{data_relatorio_str}.xlsx")
    
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
        relatorio_agregado.to_excel(writer, sheet_name='Vendas Diárias', index=False)
        df_vendas_dia.to_excel(writer, sheet_name='Detalhes', index=False)

    formatar_planilha_excel(caminho_arquivo)
    adicionar_assinatura_excel(caminho_arquivo)

    return send_file(caminho_arquivo, as_attachment=True, download_name=os.path.basename(caminho_arquivo))

if __name__ == '__main__':
    # A API será executada no endereço http://127.0.0.1:5000/
    # O modo debug reinicia a API automaticamente a cada alteração no código.
    app.run(debug=True)