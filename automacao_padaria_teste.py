import datetime
import pandas as pd
import json
import locale
import os
import time
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from collections import defaultdict
import smtplib
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
import uuid

# --- CÓDIGO DA API FLASK ---
from flask import Flask, jsonify, request
from flask_cors import CORS

# --- Códigos ANSI para cores ---
AMARELO = '\033[93m'
AZUL = '\033[94m'
VERDE = '\033[92m'
VERMELHO = '\033[91m'
RESET = '\033[0m'
NEGRITO = '\033[1m'

# Configura o locale para português do Brasil
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')

# Diretório para salvar os arquivos e subpastas
DIRETORIO_DADOS = "relatorios_padaria"
ASSINATURA = "Sistema desenvolvido por ROBSON ALVES"

# --- Configurações do e-mail (ATUALIZAR COM SUAS INFORMAÇÕES) ---
EMAIL_REMETENTE = 'robtechservice@outlook.com'
SENHA_APP = 'ioohmnnkugrsulss'
SENHA_MASTER = '120724'

# ----------------------------------------
# --- FUNÇÕES DE LÓGICA DO NEGÓCIO ---
# ----------------------------------------
def formatar_texto(texto, cor=RESET, estilo=RESET):
    """Função centralizada para formatar texto com cores e estilos."""
    return f"{estilo}{cor}{texto}{RESET}"

def verificar_diretorio():
    """Cria o diretório de dados se ele não existir."""
    if not os.path.exists(DIRETORIO_DADOS):
        os.makedirs(DIRETORIO_DADOS)

def carregar_dados():
    """Carrega os dados de um arquivo JSON. Se não existir, cria um com dados de exemplo."""
    verificar_diretorio()
    caminho_arquivo = os.path.join(DIRETORIO_DADOS, 'dados_padaria.json')
    try:
        with open(caminho_arquivo, 'r') as f:
            dados = json.load(f)
            return dados['receitas'], dados['despesas'], dados['estoque']
    except FileNotFoundError:
        print(formatar_texto("Arquivo de dados não encontrado. Criando um novo com dados de exemplo...", cor=AMARELO))
        receitas_exemplo = []
        despesas_exemplo = []
        estoque_exemplo = {
            '123456789': {'descricao': 'pão francês', 'quantidade': 100, 'valor_unitario': 0.50, 'categoria': 'Pães'},
            '987654321': {'descricao': 'pão de queijo', 'quantidade': 50, 'valor_unitario': 3.00, 'categoria': 'Pães'},
            '456789123': {'descricao': 'bolo de chocolate', 'quantidade': 10, 'valor_unitario': 25.00, 'categoria': 'Doces'},
            '789123456': {'descricao': 'torta de limão', 'quantidade': 8, 'valor_unitario': 30.00, 'categoria': 'Doces'},
            '111222333': {'descricao': 'café expresso', 'quantidade': 100, 'valor_unitario': 4.50, 'categoria': 'Bebidas'},
            '444555666': {'descricao': 'suco de laranja', 'quantidade': 20, 'valor_unitario': 6.00, 'categoria': 'Bebidas'}
        }
        salvar_dados(receitas_exemplo, despesas_exemplo, estoque_exemplo)
        return receitas_exemplo, despesas_exemplo, estoque_exemplo

def salvar_dados(receitas, despesas, estoque):
    """Salva os dados em um arquivo JSON e exibe uma mensagem de sucesso."""
    verificar_diretorio()
    caminho_arquivo = os.path.join(DIRETORIO_DADOS, 'dados_padaria.json')
    dados = {
        'receitas': receitas,
        'despesas': despesas,
        'estoque': estoque
    }
    try:
        with open(caminho_arquivo, 'w') as f:
            json.dump(dados, f, indent=4)
        print(formatar_texto("Dados salvos com sucesso!", cor=VERDE))
    except Exception as e:
        print(formatar_texto(f"Erro ao salvar os dados: {e}", cor=VERMELHO))

def adicionar_produto_estoque(estoque):
    """Permite adicionar ou atualizar um produto no estoque via terminal."""
    codigo = input("Digite o código do produto (código de barras): ")
    descricao = input("Digite a descrição do produto: ")
    try:
        quantidade = int(input("Digite a quantidade a ser adicionada: "))
        valor_unitario = float(input("Digite o valor unitário: "))
    except ValueError:
        print(formatar_texto("Entrada inválida para quantidade ou valor. Tente novamente.", cor=VERMELHO))
        return
    categoria = input("Digite a categoria do produto: ")

    if codigo in estoque:
        estoque[codigo]['quantidade'] += quantidade
        print(formatar_texto("Produto atualizado com sucesso!", cor=VERDE))
    else:
        estoque[codigo] = {
            'descricao': descricao.lower(),
            'quantidade': quantidade,
            'valor_unitario': valor_unitario,
            'categoria': categoria
        }
        print(formatar_texto("Produto adicionado com sucesso!", cor=VERDE))

def alterar_produto_estoque(estoque):
    """Permite alterar a quantidade e o valor de um produto no estoque via terminal."""
    codigo = input("Digite o código do produto que deseja alterar: ")
    if codigo not in estoque:
        print(formatar_texto("Produto não encontrado no estoque.", cor=VERMELHO))
        return
    
    try:
        nova_quantidade = int(input(f"Digite a nova quantidade para '{estoque[codigo]['descricao']}' (atual: {estoque[codigo]['quantidade']}): "))
        novo_valor = float(input(f"Digite o novo valor unitário para '{estoque[codigo]['descricao']}' (atual: R$ {estoque[codigo]['valor_unitario']:.2f}): "))
    except ValueError:
        print(formatar_texto("Entrada inválida para quantidade ou valor. Tente novamente.", cor=VERMELHO))
        return
    
    estoque[codigo]['quantidade'] = nova_quantidade
    estoque[codigo]['valor_unitario'] = novo_valor
    print(formatar_texto("Produto alterado com sucesso!", cor=VERDE))

def registrar_venda_direta(receitas, estoque):
    """Registra uma venda direta a partir do terminal."""
    codigo = input("Digite o código do produto vendido: ")
    try:
        quantidade = int(input("Digite a quantidade vendida: "))
    except ValueError:
        print(formatar_texto("Quantidade inválida. Tente novamente.", cor=VERMELHO))
        return

    if codigo not in estoque or estoque[codigo]['quantidade'] < quantidade:
        print(formatar_texto("Estoque insuficiente ou produto não encontrado.", cor=VERMELHO))
        return

    forma_pagamento = input("Digite a forma de pagamento (ex: 'Cartão', 'Dinheiro', 'PIX'): ")
    
    valor_unitario = estoque[codigo]['valor_unitario']
    valor_total = quantidade * valor_unitario
    
    estoque[codigo]['quantidade'] -= quantidade
    
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'id_transacao': uuid.uuid4().hex,
        'itens': [{
            'codigo_produto': codigo,
            'quantidade': quantidade,
            'valor_unitario': valor_unitario,
            'descricao': estoque[codigo]['descricao']
        }],
        'descricao': f"Venda de {quantidade} und. de {estoque[codigo]['descricao']}",
        'valor': valor_total,
        'data': agora,
        'tipo': 'receita',
        'forma_pagamento': forma_pagamento
    })
    
    print(formatar_texto(f"Venda de {valor_total:,.2f} registrada com sucesso!", cor=VERDE))

def registrar_venda_delivery(receitas, estoque):
    """Registra uma venda por delivery a partir do terminal."""
    codigo = input("Digite o código do produto vendido: ")
    try:
        quantidade = int(input("Digite a quantidade vendida: "))
    except ValueError:
        print(formatar_texto("Quantidade inválida. Tente novamente.", cor=VERMELHO))
        return

    if codigo not in estoque or estoque[codigo]['quantidade'] < quantidade:
        print(formatar_texto("Estoque insuficiente ou produto não encontrado.", cor=VERMELHO))
        return
        
    plataforma = input("Digite a plataforma de delivery (ex: 'iFood', 'Rappi'): ")
    
    valor_total = 0 # Valor para delivery é zero
    
    estoque[codigo]['quantidade'] -= quantidade
    
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'id_transacao': uuid.uuid4().hex,
        'itens': [{
            'codigo_produto': codigo,
            'quantidade': quantidade,
            'valor_unitario': estoque[codigo]['valor_unitario'],
            'descricao': estoque[codigo]['descricao']
        }],
        'descricao': f"Venda de {quantidade} und. de {estoque[codigo]['descricao']} (Delivery) - SEM VALOR",
        'valor': valor_total,
        'data': agora,
        'tipo': 'receita',
        'plataforma_delivery': plataforma
    })
    
    print(formatar_texto(f"Venda por delivery de {valor_total:,.2f} registrada com sucesso!", cor=VERDE))

def adicionar_receita(receitas):
    """Adiciona uma receita manual ao fluxo de caixa."""
    descricao = input("Descrição da receita: ")
    try:
        valor = float(input("Valor da receita: "))
    except ValueError:
        print(formatar_texto("Valor inválido. Tente novamente.", cor=VERMELHO))
        return

    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'descricao': descricao,
        'valor': valor,
        'data': agora,
        'tipo': 'receita',
        'forma_pagamento': 'Manual'
    })
    print(formatar_texto("Receita adicionada com sucesso!", cor=VERDE))

def adicionar_despesa(despesas):
    """Adiciona uma despesa manual ao fluxo de caixa."""
    descricao = input("Descrição da despesa: ")
    try:
        valor = float(input("Valor da despesa: "))
    except ValueError:
        print(formatar_texto("Valor inválido. Tente novamente.", cor=VERMELHO))
        return

    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    despesas.append({
        'descricao': descricao,
        'valor': valor,
        'data': agora,
        'tipo': 'despesa'
    })
    print(formatar_texto("Despesa adicionada com sucesso!", cor=VERDE))

def verificar_e_enviar_alerta_estoque(estoque, limite=10):
    """Verifica produtos com estoque baixo e envia um e-mail de alerta."""
    produtos_estoque_baixo = {
        codigo: dados['quantidade']
        for codigo, dados in estoque.items()
        if dados['quantidade'] < limite
    }

    if not produtos_estoque_baixo:
        print(formatar_texto("Nenhum produto com estoque baixo. Tudo certo!", cor=VERDE))
        return

    # Mensagem de e-mail
    assunto = "ALERTA DE ESTOQUE BAIXO"
    corpo_email = "Os seguintes produtos estão com estoque abaixo do limite:\n\n"
    for codigo, quantidade in produtos_estoque_baixo.items():
        corpo_email += f"Cód: {codigo} - Descrição: {estoque[codigo]['descricao'].capitalize()} - Quantidade: {quantidade}\n"

    corpo_email += f"\n\n---\n{ASSINATURA}"

    # Anexa o relatório completo em Excel
    caminho_relatorio = gerar_relatorio_estoque(estoque)

    # Envio do e-mail
    try:
        server = smtplib.SMTP('smtp.office365.com', 587)
        server.starttls()
        server.login(EMAIL_REMETENTE, SENHA_APP)

        msg = MIMEMultipart()
        msg['From'] = EMAIL_REMETENTE
        msg['To'] = EMAIL_REMETENTE  # Envia para si mesmo
        msg['Subject'] = assunto

        msg.attach(MIMEText(corpo_email, 'plain'))

        # Abre o arquivo em modo binário
        with open(caminho_relatorio, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition',
                            f"attachment; filename= {os.path.basename(caminho_relatorio)}")
            msg.attach(part)

        server.sendmail(EMAIL_REMETENTE, EMAIL_REMETENTE, msg.as_string())
        server.quit()
        print(formatar_texto("Alerta de estoque enviado com sucesso por e-mail!", cor=VERDE))
    except Exception as e:
        print(formatar_texto(f"Erro ao enviar o e-mail: {e}", cor=VERMELHO))
    
def limpar_tela():
    """Limpa o terminal."""
    os.system('cls' if os.name == 'nt' else 'clear')

def menu_gerar_relatorio(receitas, despesas, estoque):
    """Sub-menu para geração de relatórios."""
    while True:
        limpar_tela()
        print(formatar_texto(f"{NEGRITO}--- Menu de Relatórios ---{RESET}", cor=AZUL))
        print("1. Gerar Relatório de Receitas")
        print("2. Gerar Relatório de Despesas")
        print("3. Gerar Relatório de Estoque")
        print("4. Voltar ao Menu Principal")
        escolha = input("Escolha uma opção: ")

        if escolha == '1':
            gerar_relatorio_receitas(receitas)
        elif escolha == '2':
            gerar_relatorio_despesas(despesas)
        elif escolha == '3':
            gerar_relatorio_estoque(estoque)
        elif escolha == '4':
            break
        else:
            print(formatar_texto("Opção inválida. Tente novamente.", cor=VERMELHO))
        
        input(formatar_texto("\nPressione Enter para continuar...", cor=AZUL, estilo=NEGRITO))

def gerar_relatorio_receitas(receitas):
    """Gera e salva um relatório em Excel das receitas."""
    if not receitas:
        print(formatar_texto("Nenhuma receita para gerar relatório.", cor=AMARELO))
        return

    df = pd.DataFrame(receitas)
    df['data'] = pd.to_datetime(df['data'], format="%d/%m/%Y %H:%M:%S")
    df['valor'] = pd.to_numeric(df['valor'])

    caminho = os.path.join(DIRETORIO_DADOS, f"relatorio_receitas_{datetime.datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx")

    try:
        with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Receitas')
            workbook = writer.book
            worksheet = writer.sheets['Receitas']
            
            # Formatação
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column # Get the column index
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
            
            # Adiciona título e assinatura
            worksheet.insert_rows(1)
            worksheet.cell(row=1, column=1, value="Relatório de Receitas").font = Font(bold=True, size=14)
            worksheet.cell(row=1, column=2, value=f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font = Font(italic=True, size=10)
            
            # Formata a coluna de valores para moeda
            for cell in worksheet['C']: # Coluna 'valor'
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"R$"#,##0.00'
            
            # Adiciona o total
            total_receitas = df['valor'].sum()
            worksheet.cell(row=len(df)+4, column=1, value="Total Receitas:").font = Font(bold=True)
            total_cell = worksheet.cell(row=len(df)+4, column=2, value=total_receitas)
            total_cell.number_format = '"R$"#,##0.00'
            total_cell.font = Font(bold=True)
            
            print(formatar_texto(f"Relatório de receitas gerado com sucesso em '{caminho}'", cor=VERDE))
    except Exception as e:
        print(formatar_texto(f"Erro ao gerar relatório de receitas: {e}", cor=VERMELHO))

def gerar_relatorio_despesas(despesas):
    """Gera e salva um relatório em Excel das despesas."""
    if not despesas:
        print(formatar_texto("Nenhuma despesa para gerar relatório.", cor=AMARELO))
        return

    df = pd.DataFrame(despesas)
    df['data'] = pd.to_datetime(df['data'], format="%d/%m/%Y %H:%M:%S")
    df['valor'] = pd.to_numeric(df['valor'])

    caminho = os.path.join(DIRETORIO_DADOS, f"relatorio_despesas_{datetime.datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx")
    
    try:
        with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Despesas')
            workbook = writer.book
            worksheet = writer.sheets['Despesas']
            
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
            
            worksheet.insert_rows(1)
            worksheet.cell(row=1, column=1, value="Relatório de Despesas").font = Font(bold=True, size=14)
            worksheet.cell(row=1, column=2, value=f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font = Font(italic=True, size=10)
            
            for cell in worksheet['C']:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"R$"#,##0.00'
            
            total_despesas = df['valor'].sum()
            worksheet.cell(row=len(df)+4, column=1, value="Total Despesas:").font = Font(bold=True)
            total_cell = worksheet.cell(row=len(df)+4, column=2, value=total_despesas)
            total_cell.number_format = '"R$"#,##0.00'
            total_cell.font = Font(bold=True)

            print(formatar_texto(f"Relatório de despesas gerado com sucesso em '{caminho}'", cor=VERDE))
    except Exception as e:
        print(formatar_texto(f"Erro ao gerar relatório de despesas: {e}", cor=VERMELHO))

def gerar_relatorio_estoque(estoque):
    """Gera e salva um relatório em Excel do estoque. Retorna o caminho do arquivo."""
    if not estoque:
        print(formatar_texto("Nenhum produto no estoque para gerar relatório.", cor=AMARELO))
        return None

    df_data = [
        {
            'Código': cod,
            'Descrição': dados['descricao'].capitalize(),
            'Quantidade': dados['quantidade'],
            'Valor Unitário': dados['valor_unitario'],
            'Categoria': dados['categoria']
        }
        for cod, dados in estoque.items()
    ]
    df = pd.DataFrame(df_data)

    caminho = os.path.join(DIRETORIO_DADOS, f"relatorio_estoque_{datetime.datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx")

    try:
        with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Estoque')
            workbook = writer.book
            worksheet = writer.sheets['Estoque']

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
            
            worksheet.insert_rows(1)
            worksheet.cell(row=1, column=1, value="Relatório de Estoque").font = Font(bold=True, size=14)
            worksheet.cell(row=1, column=2, value=f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font = Font(italic=True, size=10)
            
            for cell in worksheet['D']: # Coluna 'Valor Unitário'
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"R$"#,##0.00'
                    
            print(formatar_texto(f"Relatório de estoque gerado com sucesso em '{caminho}'", cor=VERDE))
            return caminho
    except Exception as e:
        print(formatar_texto(f"Erro ao gerar relatório de estoque: {e}", cor=VERMELHO))
        return None

def menu_principal():
    """Exibe o menu principal e processa as opções."""
    receitas, despesas, estoque = carregar_dados()
    while True:
        limpar_tela()
        print(formatar_texto(f"{NEGRITO}--- Menu Principal ---{RESET}", cor=AZUL))
        print("1. Gerar Relatórios")
        print("2. Adicionar Receita")
        print("3. Adicionar Despesa")
        print("4. Adicionar Produto ao Estoque")
        print("5. Alterar Produto no Estoque")
        print("6. Registrar Venda Direta")
        print("7. Registrar Venda Delivery")
        print("8. Verificar Estoque e Enviar Alerta")
        print(formatar_texto("9. Sair", cor=VERMELHO, estilo=NEGRITO))
        
        escolha = input("Escolha uma opção: ")

        if escolha == '1':
            menu_gerar_relatorio(receitas, despesas, estoque)
        elif escolha == '2':
            adicionar_receita(receitas)
        elif escolha == '3':
            adicionar_despesa(despesas)
        elif escolha == '4':
            adicionar_produto_estoque(estoque)
        elif escolha == '5':
            alterar_produto_estoque(estoque)
        elif escolha == '6':
            registrar_venda_direta(receitas, estoque)
        elif escolha == '7':
            registrar_venda_delivery(receitas, estoque)
        elif escolha == '8':
            verificar_e_enviar_alerta_estoque(estoque)
        elif escolha == '9':
            print("Salvando dados...")
            salvar_dados(receitas, despesas, estoque)
            print("Encerrando o programa.")
            break
        else:
            print(formatar_texto("Opção inválida. Tente novamente.", cor=VERMELHO))
        
        input(formatar_texto("\nPressione Enter para continuar...", cor=AZUL, estilo=NEGRITO))
        limpar_tela()

# ----------------------------------------
# --- FUNÇÕES DE API (FLASK) ---
# ----------------------------------------
app = Flask(__name__)
CORS(app)

@app.route('/estoque', methods=['GET'])
def get_estoque():
    """Endpoint para a API web buscar o estoque atual."""
    receitas, despesas, estoque = carregar_dados()
    return jsonify(estoque)

@app.route('/estoque/add', methods=['POST'])
def add_produto():
    """Endpoint para adicionar/atualizar um produto via API."""
    data = request.json
    codigo = data.get('codigo')
    descricao = data.get('descricao')
    quantidade = data.get('quantidade')
    valor_unitario = data.get('valor_unitario')
    categoria = data.get('categoria')

    if not all([codigo, descricao, quantidade, valor_unitario]):
        return jsonify({"error": "Dados incompletos"}), 400

    receitas, despesas, estoque = carregar_dados()
    if codigo in estoque:
        estoque[codigo]['quantidade'] += quantidade
        estoque[codigo]['valor_unitario'] = valor_unitario
        estoque[codigo]['categoria'] = categoria
        message = "Produto atualizado com sucesso!"
    else:
        estoque[codigo] = {
            'descricao': descricao.lower(),
            'quantidade': quantidade,
            'valor_unitario': valor_unitario,
            'categoria': categoria
        }
        message = "Produto adicionado com sucesso!"

    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": message})

@app.route('/estoque/alterar', methods=['POST'])
def alterar_produto_api():
    """Endpoint para alterar a quantidade e o valor de um produto via API."""
    data = request.json
    codigo = data.get('codigo')
    nova_quantidade = data.get('quantidade')
    novo_valor = data.get('valor')

    if not all([codigo, nova_quantidade, novo_valor]):
        return jsonify({"error": "Dados incompletos"}), 400

    receitas, despesas, estoque = carregar_dados()
    if codigo not in estoque:
        return jsonify({"error": "Produto não encontrado no estoque."}), 404

    try:
        estoque[codigo]['quantidade'] = int(nova_quantidade)
        estoque[codigo]['valor_unitario'] = float(novo_valor)
        salvar_dados(receitas, despesas, estoque)
        return jsonify({"message": "Produto alterado com sucesso!"})
    except (ValueError, TypeError):
        return jsonify({"error": "Dados de quantidade ou valor inválidos."}), 400

@app.route('/vendas/direta', methods=['POST'])
def post_venda_direta():
    """ Endpoint para registrar uma venda direta com múltiplos itens via API. O payload deve conter a lista de itens do carrinho. """
    data = request.json
    carrinho = data.get('carrinho')
    forma_pagamento = data.get('forma_pagamento')

    if not carrinho or not forma_pagamento:
        return jsonify({"error": "Dados de venda incompletos"}), 400

    receitas, despesas, estoque = carregar_dados()
    
    valor_total = 0
    itens_vendidos = []

    # 1. Verifica o estoque de todos os itens antes de processar a venda
    for item in carrinho:
        codigo = item.get('codigo')
        quantidade = item.get('quantidade')
        
        if codigo not in estoque or estoque[codigo]['quantidade'] < quantidade:
            return jsonify({"error": f"Estoque insuficiente para o item {codigo} ou produto não encontrado."}), 400
            
        valor_unitario = estoque[codigo]['valor_unitario']
        valor_total += quantidade * valor_unitario

        itens_vendidos.append({
            'codigo_produto': codigo,
            'quantidade': quantidade,
            'valor_unitario': valor_unitario,
            'descricao': estoque[codigo]['descricao']
        })

    # 2. Processa a venda e atualiza o estoque
    for item in carrinho:
        codigo = item.get('codigo')
        quantidade = item.get('quantidade')
        estoque[codigo]['quantidade'] -= quantidade
        
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'id_transacao': uuid.uuid4().hex,
        'itens': itens_vendidos,
        'descricao': f"Venda de {len(carrinho)} itens",
        'valor': valor_total,
        'data': agora,
        'tipo': 'receita',
        'forma_pagamento': forma_pagamento
    })

    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": f"Venda de R$ {valor_total:,.2f} registrada com sucesso!"})

@app.route('/vendas/delivery', methods=['POST'])
def post_venda_delivery():
    """Endpoint para registrar uma venda de delivery com múltiplos itens via API."""
    data = request.json
    carrinho = data.get('carrinho')
    plataforma = data.get('plataforma')

    if not carrinho or not plataforma:
        return jsonify({"error": "Dados de venda incompletos"}), 400

    receitas, despesas, estoque = carregar_dados()

    # 1. Verifica o estoque de todos os itens antes de processar a venda
    for item in carrinho:
        codigo = item.get('codigo')
        quantidade = item.get('quantidade')
        if codigo not in estoque or estoque[codigo]['quantidade'] < quantidade:
            return jsonify({"error": f"Estoque insuficiente para o item {codigo} ou produto não encontrado."}), 400

    itens_vendidos = []
    for item in carrinho:
        codigo = item.get('codigo')
        quantidade = item.get('quantidade')
        estoque[codigo]['quantidade'] -= quantidade
        
        itens_vendidos.append({
            'codigo_produto': codigo,
            'quantidade': quantidade,
            'valor_unitario': estoque[codigo]['valor_unitario'],
            'descricao': estoque[codigo]['descricao']
        })
        
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'id_transacao': uuid.uuid4().hex,
        'itens': itens_vendidos,
        'descricao': f"Venda de {len(carrinho)} itens (Delivery)",
        'valor': 0, # Venda delivery sem valor de transação
        'data': agora,
        'tipo': 'receita',
        'plataforma_delivery': plataforma
    })

    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": "Venda por delivery registrada com sucesso!"})

@app.route('/receitas/add', methods=['POST'])
def add_receita():
    """Endpoint para adicionar uma receita manual via API."""
    data = request.json
    descricao = data.get('descricao')
    valor = data.get('valor')
    
    if not all([descricao, valor]):
        return jsonify({"error": "Dados incompletos"}), 400

    try:
        valor = float(valor)
    except ValueError:
        return jsonify({"error": "Valor inválido"}), 400

    receitas, despesas, estoque = carregar_dados()
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    receitas.append({
        'descricao': descricao,
        'valor': valor,
        'data': agora,
        'tipo': 'receita',
        'forma_pagamento': 'Manual'
    })
    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": "Receita adicionada com sucesso!"})

@app.route('/despesas/add', methods=['POST'])
def add_despesa():
    """Endpoint para adicionar uma despesa manual via API."""
    data = request.json
    descricao = data.get('descricao')
    valor = data.get('valor')
    
    if not all([descricao, valor]):
        return jsonify({"error": "Dados incompletos"}), 400

    try:
        valor = float(valor)
    except ValueError:
        return jsonify({"error": "Valor inválido"}), 400

    receitas, despesas, estoque = carregar_dados()
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    despesas.append({
        'descricao': descricao,
        'valor': valor,
        'data': agora,
        'tipo': 'despesa'
    })
    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": "Despesa adicionada com sucesso!"})

@app.route('/vendas/diarias', methods=['GET'])
def get_vendas_diarias():
    """Endpoint para buscar todas as vendas do dia atual."""
    receitas, _, _ = carregar_dados()
    hoje = datetime.datetime.now().strftime("%d/%m/%Y")
    
    vendas_hoje = [
        venda for venda in receitas
        if 'venda' in venda.get('descricao', '').lower() and hoje in venda['data']
    ]

    return jsonify(vendas_hoje)

@app.route('/vendas/cancelar/<id_transacao>', methods=['POST'])
def cancelar_venda(id_transacao):
    """Endpoint para cancelar uma venda e reverter o estoque."""
    data = request.json
    senha = data.get('senha')
    if senha != SENHA_MASTER:
        return jsonify({"error": "Senha master incorreta."}), 403

    receitas, despesas, estoque = carregar_dados()
    
    venda_para_cancelar = None
    for venda in receitas:
        if venda.get('id_transacao') == id_transacao:
            venda_para_cancelar = venda
            break
            
    if not venda_para_cancelar:
        return jsonify({"error": "Venda não encontrada."}), 404

    if 'CANCELADA' in venda_para_cancelar.get('descricao', '').upper():
        return jsonify({"error": "Venda já está cancelada."}), 400

    # Adiciona os itens de volta ao estoque
    for item in venda_para_cancelar.get('itens', []):
        codigo = item.get('codigo_produto')
        quantidade = item.get('quantidade')
        if codigo in estoque:
            estoque[codigo]['quantidade'] += quantidade

    # Marca a venda como cancelada
    venda_para_cancelar['descricao'] += " (CANCELADA)"
    venda_para_cancelar['valor'] = 0 # Valor da venda cancelada se torna 0
    
    salvar_dados(receitas, despesas, estoque)
    return jsonify({"message": "Venda cancelada e estoque revertido com sucesso!"})

@app.route('/relatorios', methods=['POST'])
def gerar_relatorio_api():
    """Endpoint para gerar relatórios via API."""
    data = request.json
    tipo = data.get('tipo')
    
    receitas, despesas, estoque = carregar_dados()
    
    if tipo == 'receitas':
        gerar_relatorio_receitas(receitas)
        return jsonify({"message": "Relatório de receitas gerado!"})
    elif tipo == 'despesas':
        gerar_relatorio_despesas(despesas)
        return jsonify({"message": "Relatório de despesas gerado!"})
    elif tipo == 'estoque':
        gerar_relatorio_estoque(estoque)
        return jsonify({"message": "Relatório de estoque gerado!"})
    else:
        return jsonify({"error": "Tipo de relatório inválido."}), 400

@app.route('/estoque/alerta', methods=['POST'])
def enviar_alerta_estoque_api():
    """Endpoint para enviar alerta de estoque baixo via API."""
    receitas, despesas, estoque = carregar_dados()
    verificar_e_enviar_alerta_estoque(estoque)
    return jsonify({"message": "Verificação de estoque e alerta enviados."})

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == 'terminal':
        menu_principal()
    else:
        app.run(debug=True)